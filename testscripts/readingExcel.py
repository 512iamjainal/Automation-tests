import openpyxl

#File-->workbook-->sheet-->rows-->cells
file="C:\\Selenium_practice\\data.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row  #counts number of rows
columns=sheet.max_column  #counts numner of columns

#reading all rows and columns from sheet
for r in range(1,rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(r,c).value,end="     ")
    print()